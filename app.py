"""
═════════════════════════════════════════════════════════════════════════════
CRTV FLY-AWAY MONITOR v3.3 — Authentification + Rapports
═════════════════════════════════════════════════════════════════════════════
+ Authentification agents (nom, matricule, password)
+ Lieu de production
+ Génération rapports de fin de transmission (PDF/Excel)

Auteur    : Moussa Manga Asser
Version   : 3.3.0
Contact   : +237 690 537 181 | assermoussa19@gmail.com
═════════════════════════════════════════════════════════════════════════════
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
import plotly.graph_objects as go
import plotly.express as px
from PIL import Image
import warnings
import os
import json
import hashlib
from io import BytesIO
import hmac

warnings.filterwarnings('ignore')

# ════════════════════════════════════════════════════════════════════════════
# CONFIGURATION GLOBALE
# ════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="CRTV Fly-Away v3.3 — Authentification + Rapports",
    page_icon="🛰️",
    layout="wide",
    initial_sidebar_state="expanded"
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
ASSETS_DIR = os.path.join(BASE_DIR, "assets")

DATASET_FILENAME = "flyaway_log_annuel_2025-1.xlsx"
DEFAULT_EXCEL_PATH = os.path.join(DATA_DIR, DATASET_FILENAME)
ANCIEN_DATASET_PATH = os.path.join(DATA_DIR, "flyaway log annuel 2025-1.xlsx")
DEFAULT_LOGO_PATH = os.path.join(ASSETS_DIR, "logo-crtv.png")
DEFAULT_SHEET_NAME = "Donnees_2025"

FENETRE_RECENTE = timedelta(hours=1)
SEUIL_LATENCE_WARN = 720
SEUIL_LATENCE_CRIT = 800

# Base de données agents (en production : utiliser une vraie DB)
# Structure : {"matricule": {"nom": str, "password_hash": str, "lieu": str}}
AGENTS_DB_FILE = os.path.join(BASE_DIR, "agents.json")

LIEUX_DISPONIBLES = [
    "Yaoundé — Centre Principal",
    "Douala — Antenne Côtière",
    "Buea — Station Montagne",
    "Bafoussam — Antenne Ouest",
    "Garoua — Antenne Nord",
    "Bertoua — Antenne Est",
    "Kribi — Station Côte",
    "Kumba — Antenne Anglophone",
    "Autre (à préciser)"
]

# ════════════════════════════════════════════════════════════════════════════
# UTILITAIRES AUTHENTIFICATION
# ════════════════════════════════════════════════════════════════════════════

def hash_password(password: str) -> str:
    """Hash un mot de passe avec SHA-256."""
    return hashlib.sha256(password.encode()).hexdigest()


def charger_agents_db():
    """Charge la base de données des agents depuis un fichier JSON."""
    if os.path.exists(AGENTS_DB_FILE):
        try:
            with open(AGENTS_DB_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            st.error(f"Erreur lecture agents.json : {e}")
            return {}
    # Initialiser avec agents de démo
    return {
        "MAT001": {
            "nom": "Moussa Manga Asser",
            "password_hash": hash_password("demo123"),
            "lieu": "Yaoundé — Centre Principal",
            "role": "Ingénieur Senior"
        },
        "MAT002": {
            "nom": "Jean Dupont",
            "password_hash": hash_password("demo123"),
            "lieu": "Douala — Antenne Côtière",
            "role": "Technicien"
        },
        "MAT003": {
            "nom": "Marie Tagne",
            "password_hash": hash_password("demo123"),
            "lieu": "Buea — Station Montagne",
            "role": "Opérateur"
        }
    }


def sauvegarder_agents_db(agents_db):
    """Sauvegarde la base de données des agents."""
    os.makedirs(BASE_DIR, exist_ok=True)
    with open(AGENTS_DB_FILE, 'w', encoding='utf-8') as f:
        json.dump(agents_db, f, ensure_ascii=False, indent=2)


def verifier_authentification(matricule: str, password: str) -> tuple[bool, str]:
    """
    Vérifie les credentials d'un agent.
    Retourne (succès, message_erreur_ou_nom)
    """
    agents_db = charger_agents_db()
    
    if matricule not in agents_db:
        return False, "Matricule introuvable"
    
    agent = agents_db[matricule]
    password_hash = hash_password(password)
    
    if agent["password_hash"] != password_hash:
        return False, "Mot de passe incorrect"
    
    return True, agent["nom"]


# ════════════════════════════════════════════════════════════════════════════
# CACHE GLOBAL
# ════════════════════════════════════════════════════════════════════════════

@st.cache_resource
def init_global_state():
    """Initialise l'état global une seule fois."""
    return {"chat_history": []}


# ════════════════════════════════════════════════════════════════════════════
# CHARGEMENT DONNÉES
# ════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=300, show_spinner=False)
def charger_donnees_depuis_chemin(file_path: str, sheet_name: str):
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        skiprows=1,
        dtype={
            'timestamp': 'object',
            'latence_totale_ms': 'float32',
            'cn0_dbhz': 'float32',
            'esn0_db': 'float32',
            'qualite_signal_pct': 'float32',
            'jitter_ms': 'float32',
            'ber': 'float64',
            'temperature_hpa_c': 'float32',
            'temperature_buc_c': 'float32'
        }
    )
    return df


@st.cache_data(ttl=300, show_spinner=False)
def charger_donnees_depuis_upload(file_bytes: bytes, sheet_name: str):
    import io
    df = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        skiprows=1
    )
    return df


def nettoyer_donnees(df: pd.DataFrame) -> pd.DataFrame:
    """Nettoyage commun des données."""
    if df is None or df.empty:
        return df
    df['timestamp'] = pd.to_datetime(df['timestamp'], errors='coerce')
    df = df.dropna(subset=['timestamp'])
    df = df.sort_values('timestamp', ascending=False)
    return df


def charger_donnees(file_path: str, sheet_name: str, uploaded_file=None):
    """Charge les données (fichier par défaut ou uploader)."""
    try:
        if uploaded_file is not None:
            df = charger_donnees_depuis_upload(uploaded_file.getvalue(), sheet_name)
            return nettoyer_donnees(df), None

        chemins = [file_path, ANCIEN_DATASET_PATH]
        for chemin in chemins:
            if os.path.exists(chemin):
                df = charger_donnees_depuis_chemin(chemin, sheet_name)
                return nettoyer_donnees(df), None

        return None, (
            f"Fichier Excel introuvable. Vérifiez que '{DATASET_FILENAME}' "
            "est bien dans le dossier data/ du projet."
        )
    except Exception as e:
        return None, f"Erreur de chargement : {e}"


def obtenir_reference_temporelle(df: pd.DataFrame) -> datetime:
    """Référence temporelle basée sur les données, pas sur l'horloge système."""
    if df is None or df.empty:
        return datetime.now()
    return df['timestamp'].max()


# ════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION DE RAPPORTS
# ════════════════════════════════════════════════════════════════════════════

def generer_rapport_excel(
    df_recent: pd.DataFrame,
    df_complet: pd.DataFrame,
    agent_info: dict,
    reference_now: datetime,
    chatbot_responses: list = None
) -> BytesIO:
    """
    Génère un rapport Excel complet de fin de transmission.
    Retourne un BytesIO prêt à télécharger.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    
    # ── Feuille 1 : Résumé Exécutif ──────────────────────────────────────
    ws_resume = wb.active
    ws_resume.title = "Résumé"
    
    ws_resume['A1'] = "CRTV FLY-AWAY — RAPPORT DE FIN DE TRANSMISSION"
    ws_resume['A1'].font = Font(bold=True, size=14)
    ws_resume.merge_cells('A1:D1')
    
    row = 3
    ws_resume[f'A{row}'] = "Informations Transmission"
    ws_resume[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    data_resume = [
        ("Date/Heure Rapport", reference_now.strftime("%Y-%m-%d %H:%M:%S")),
        ("Agent", agent_info.get("nom", "N/A")),
        ("Matricule", agent_info.get("matricule", "N/A")),
        ("Lieu Production", agent_info.get("lieu", "N/A")),
        ("Rôle", agent_info.get("role", "N/A")),
        ("", ""),
        ("Statistiques Transmission", ""),
        ("Latence Moyenne", f"{df_recent['latence_totale_ms'].mean():.0f}ms"),
        ("Latence Min/Max", f"{df_recent['latence_totale_ms'].min():.0f} / {df_recent['latence_totale_ms'].max():.0f}ms"),
        ("Qualité Signal Moyenne", f"{df_recent['qualite_signal_pct'].mean():.1f}%" if 'qualite_signal_pct' in df_recent.columns else "N/A"),
        ("C/N₀ Moyen", f"{df_recent['cn0_dbhz'].mean():.1f}dBHz" if 'cn0_dbhz' in df_recent.columns else "N/A"),
        ("Température HPA", f"{df_recent['temperature_hpa_c'].mean():.1f}°C" if 'temperature_hpa_c' in df_recent.columns else "N/A"),
        ("", ""),
        ("Statut Transmission", ""),
        ("État Global", _determiner_etat(df_recent)),
        ("Nombre Points Mesure", len(df_recent)),
        ("Période Fenêtre Temps Réel", f"{int(FENETRE_RECENTE.total_seconds()//3600)}h"),
    ]
    
    for label, value in data_resume:
        if label and not value:
            ws_resume[f'A{row}'] = label
            ws_resume[f'A{row}'].font = Font(bold=True, size=11)
        else:
            ws_resume[f'A{row}'] = label
            ws_resume[f'B{row}'] = value
        row += 1
    
    ws_resume.column_dimensions['A'].width = 30
    ws_resume.column_dimensions['B'].width = 40
    
    # ── Feuille 2 : Données Détaillées ───────────────────────────────────
    ws_data = wb.create_sheet("Données Détaillées")
    
    for col_idx, col_name in enumerate(df_recent.columns, 1):
        ws_data.cell(row=1, column=col_idx, value=col_name)
        ws_data.cell(row=1, column=col_idx).font = Font(bold=True)
    
    for row_idx, row_data in enumerate(df_recent.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    for col in ws_data.columns:
        max_length = 15
        column = get_column_letter(col[0].column)
        ws_data.column_dimensions[column].width = max_length
    
    # ── Feuille 3 : Analyses ─────────────────────────────────────────────
    ws_analyse = wb.create_sheet("Analyses")
    
    row = 1
    ws_analyse[f'A{row}'] = "ANALYSES STATISTIQUES"
    ws_analyse[f'A{row}'].font = Font(bold=True, size=12)
    row += 2
    
    analyses = [
        ("Latence", {
            "Moyenne": f"{df_recent['latence_totale_ms'].mean():.0f}ms",
            "Médiane": f"{df_recent['latence_totale_ms'].median():.0f}ms",
            "Écart-type": f"{df_recent['latence_totale_ms'].std():.0f}ms",
            "Min": f"{df_recent['latence_totale_ms'].min():.0f}ms",
            "Max": f"{df_recent['latence_totale_ms'].max():.0f}ms"
        }),
    ]
    
    if 'qualite_signal_pct' in df_recent.columns:
        analyses.append(("Qualité Signal", {
            "Moyenne": f"{df_recent['qualite_signal_pct'].mean():.1f}%",
            "Min": f"{df_recent['qualite_signal_pct'].min():.1f}%",
            "Max": f"{df_recent['qualite_signal_pct'].max():.1f}%"
        }))
    
    if 'temperature_hpa_c' in df_recent.columns:
        analyses.append(("Température HPA", {
            "Moyenne": f"{df_recent['temperature_hpa_c'].mean():.1f}°C",
            "Min": f"{df_recent['temperature_hpa_c'].min():.1f}°C",
            "Max": f"{df_recent['temperature_hpa_c'].max():.1f}°C"
        }))
    
    for analyse_nom, stats in analyses:
        ws_analyse[f'A{row}'] = analyse_nom
        ws_analyse[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        for stat_label, stat_value in stats.items():
            ws_analyse[f'A{row}'] = stat_label
            ws_analyse[f'B{row}'] = stat_value
            row += 1
        row += 1
    
    ws_analyse.column_dimensions['A'].width = 25
    ws_analyse.column_dimensions['B'].width = 20
    
    # ── Feuille 4 : Recommandations ──────────────────────────────────────
    ws_recom = wb.create_sheet("Recommandations")
    
    row = 1
    ws_recom[f'A{row}'] = "RECOMMANDATIONS ET ACTIONS"
    ws_recom[f'A{row}'].font = Font(bold=True, size=12)
    row += 2
    
    lat_mean = df_recent['latence_totale_ms'].mean()
    
    recommendations = _generer_recommandations_texte(df_recent, lat_mean)
    
    ws_recom[f'A{row}'] = "Basé sur l'analyse actuelle :"
    row += 1
    for line in recommendations.split('\n'):
        if line.strip():
            ws_recom[f'A{row}'] = line
            row += 1
        else:
            row += 1
    
    ws_recom.column_dimensions['A'].width = 80
    ws_recom.wrap_text = True
    
    # Sérialiser
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _determiner_etat(df_recent: pd.DataFrame) -> str:
    """Détermine l'état global du système."""
    if df_recent.empty:
        return "Pas de données"
    lat = df_recent['latence_totale_ms'].mean()
    qual = df_recent['qualite_signal_pct'].mean() if 'qualite_signal_pct' in df_recent.columns else 100
    
    if lat > SEUIL_LATENCE_CRIT:
        return "🔴 CRITIQUE"
    elif lat > SEUIL_LATENCE_WARN:
        return "🟡 DÉGRADÉ"
    else:
        return "🟢 NORMAL"


def _generer_recommandations_texte(df_recent: pd.DataFrame, lat_mean: float) -> str:
    """Génère les recommandations en texte brut pour le rapport."""
    recom = ""
    
    if lat_mean > SEUIL_LATENCE_CRIT:
        recom += "🔴 LATENCE CRITIQUE (> 800ms) — INTERVENTION IMMÉDIATE REQUISE\n"
        recom += "  1. Vérifier pointage antenne (Azimut/Élévation)\n"
        recom += "  2. Consulter météo locale (pluie, vent)\n"
        recom += "  3. Augmenter FEC DVB-S2 si C/N₀ bas\n"
        recom += "  4. Préparer basculement vers liaison alternative\n"
    elif lat_mean > SEUIL_LATENCE_WARN:
        recom += "🟡 LATENCE ÉLEVÉE (> 720ms) — À SURVEILLER\n"
        recom += "  1. Optimiser paramètres DVB-S2\n"
        recom += "  2. Ajuster buffer jitter encodeur\n"
        recom += "  3. Surveiller température HPA\n"
    else:
        recom += "🟢 SITUATION NORMALE\n"
        recom += "  • Maintenir configuration actuelle\n"
        recom += "  • Surveiller tendances horaires\n"
    
    temp = df_recent['temperature_hpa_c'].mean() if 'temperature_hpa_c' in df_recent.columns else 0
    if temp > 65:
        recom += f"\n⚠️  Température HPA élevée ({temp:.1f}°C)\n"
        recom += "  → Réduire puissance HPA ou améliorer refroidissement\n"
    
    if 'attenuation_pluie_db' in df_recent.columns:
        pluie = df_recent['attenuation_pluie_db'].mean()
        if pluie > 2:
            recom += f"\n🌧️  Affaiblissement pluie important ({pluie:.2f}dB)\n"
            recom += "  → Augmenter FEC DVB-S2, réduire symbol rate\n"
    
    return recom


# ════════════════════════════════════════════════════════════════════════════
# MODÈLE IA
# ════════════════════════════════════════════════════════════════════════════

class ModeleAIOptimisee:
    """Modèle IA pour prédictions."""

    def __init__(self):
        self.model = RandomForestRegressor(
            n_estimators=50,
            max_depth=10,
            n_jobs=-1,
            random_state=42
        )
        self.scaler = StandardScaler()
        self.trained = False
        self.features = []

    def entrainer_rapide(self, df):
        try:
            params_cles = [
                'temperature_hpa_c', 'temperature_buc_c',
                'attenuation_pluie_db', 'symbol_rate_msps'
            ]
            cols_dispo = [p for p in params_cles if p in df.columns]
            if not cols_dispo or 'latence_totale_ms' not in df.columns:
                return False

            X = df[cols_dispo].dropna()
            y = df.loc[X.index, 'latence_totale_ms']

            if len(X) > 50:
                self.features = X.columns.tolist()
                X_scaled = self.scaler.fit_transform(X)
                self.model.fit(X_scaled, y)
                self.trained = True
                return True
        except Exception:
            pass
        return False

    def predire_rapide(self, row_dict):
        if not self.trained:
            return None
        try:
            X_input = np.array([[row_dict.get(f, 0) for f in self.features]])
            X_scaled = self.scaler.transform(X_input)
            return float(self.model.predict(X_scaled)[0])
        except Exception:
            return None


# ════════════════════════════════════════════════════════════════════════════
# CHATBOT IA
# ════════════════════════════════════════════════════════════════════════════

class ChatbotIAPrediction:
    """Chatbot pour prédictions et anomalies."""

    def __init__(self, model_ia, df_recent, df_complet):
        self.model = model_ia
        self.df_recent = df_recent
        self.df_complet = df_complet

    def repondre(self, question):
        q = question.lower()

        if any(w in q for w in ['prédi', 'futur', 'demain', 'dans 15']):
            return self._expliquer_predictions()
        elif any(w in q for w in ['latence', 'délai', 'delay']):
            return self._analyser_latence()
        elif any(w in q for w in ['qualité', 'signal', 'quality']):
            return self._analyser_qualite()
        elif any(w in q for w in ['temp', 'température', 'heat']):
            return self._analyser_temperature()
        elif any(w in q for w in ['pluie', 'météo', 'weather', 'rain']):
            return self._analyser_pluie()
        elif any(w in q for w in ['recommand', 'action', 'faire', 'quoi faire']):
            return self._recommandations_actions()
        elif any(w in q for w in ['comment', 'pourquoi', 'quoi', 'status']):
            return self._resume_global()
        else:
            return self._reponse_defaut(question)

    def _expliquer_predictions(self):
        if self.df_recent.empty:
            return "❌ Pas assez de données pour les prédictions."

        latence_actuelle = self.df_recent['latence_totale_ms'].mean()
        latences_recent = self.df_recent['latence_totale_ms'].tail(10).values

        if len(latences_recent) > 2:
            trend = np.polyfit(range(len(latences_recent)), latences_recent, 1)[0]
            latence_pred = latence_actuelle + trend * 3
        else:
            latence_pred = latence_actuelle

        confiance = (
            "⭐⭐⭐⭐ Très haute (historique stable)" if np.std(latences_recent) < 30
            else "⭐⭐⭐ Bonne (variations acceptables)" if np.std(latences_recent) < 50
            else "⭐⭐ Modérée (variations importantes)"
        )

        if latence_pred > SEUIL_LATENCE_CRIT:
            interp = "🔴 **CRITIQUE** : Action urgente requise MAINTENANT"
        elif latence_pred > SEUIL_LATENCE_WARN:
            interp = "🟡 **ATTENTION** : À surveiller attentivement"
        else:
            interp = "🟢 **OK** : Situation stable"

        return f"""
### 🔮 Prédictions — 15 prochaines minutes

**Latence actuelle** : {latence_actuelle:.0f}ms  
**Latence prédite** : {latence_pred:.0f}ms  
**Tendance** : {"📈 HAUSSE" if latence_pred > latence_actuelle else "📉 BAISSE"}  
**Confiance** : {confiance}

{interp}
"""

    def _analyser_latence(self):
        if self.df_recent.empty:
            return "❌ Pas de données latence."
        lat = self.df_recent['latence_totale_ms']
        statut = (
            "🔴 CRITIQUE - Intervention urgente" if lat.mean() > SEUIL_LATENCE_CRIT
            else "🟡 DÉGRADÉE - À surveiller" if lat.mean() > SEUIL_LATENCE_WARN
            else "🟢 NORMALE - Aucune action"
        )
        return f"""
### ⏱️ Analyse Latence

**Actuelle** : {lat.iloc[0]:.0f}ms | **Moyenne** : {lat.mean():.0f}ms  
**Min/Max** : {lat.min():.0f}ms / {lat.max():.0f}ms | **Écart-type** : {lat.std():.0f}ms  
**Statut** : {statut}
"""

    def _analyser_qualite(self):
        if self.df_recent.empty or 'qualite_signal_pct' not in self.df_recent.columns:
            return "❌ Pas de données qualité."
        qual = self.df_recent['qualite_signal_pct'].mean()
        cn0 = self.df_recent['cn0_dbhz'].mean() if 'cn0_dbhz' in self.df_recent.columns else np.nan
        statut = (
            "🔴 Insuffisante" if qual < 80
            else "🟡 Dégradée" if qual < 85
            else "🟢 Excellente"
        )
        return f"### 📡 Qualité Signal\n\n**Signal** : {qual:.1f}% ({statut}) | **C/N₀** : {cn0:.1f}dBHz"

    def _analyser_temperature(self):
        if 'temperature_hpa_c' not in self.df_recent.columns or self.df_recent.empty:
            return "❌ Données température non disponibles."
        temp = self.df_recent['temperature_hpa_c'].mean()
        statut = (
            "🛑 CRITIQUE" if temp > 75 else "🔴 Élevée" if temp > 65
            else "🟡 Modérée" if temp > 50 else "🟢 Optimale"
        )
        return f"### 🌡️ Température HPA\n\n**Temp** : {temp:.1f}°C ({statut})"

    def _analyser_pluie(self):
        if 'attenuation_pluie_db' not in self.df_recent.columns or self.df_recent.empty:
            return "❌ Données pluie non disponibles."
        pluie = self.df_recent['attenuation_pluie_db'].mean()
        return f"### 🌧️ Impact Pluie\n\n**Affaiblissement** : {pluie:.2f}dB (0-0.5: OK | 0.5-2: léger | 2-4: modéré | >4: fort)"

    def _resume_global(self):
        if self.df_recent.empty:
            return "❌ Pas de données."
        lat = self.df_recent['latence_totale_ms'].mean()
        etat = "🟢 NORMAL" if lat < SEUIL_LATENCE_WARN else \
               "🟡 DÉGRADÉ" if lat < SEUIL_LATENCE_CRIT else "🔴 CRITIQUE"
        return f"### 📊 Statut\n\n**Latence** : {lat:.0f}ms | **État** : {etat}"

    def _recommandations_actions(self):
        lat = self.df_recent['latence_totale_ms'].mean() if not self.df_recent.empty else 0
        return f"""
### ⚡ Actions Recommandées

**Immédiate** (<5min) : vérifier pointage, consulter météo, augmenter FEC DVB-S2  
**Court terme** (<30min) : optimiser paramètres DVB-S2, ajuster buffer jitter  
**Moyen terme** (<2h) : générer rapports, informer direction si critique  
**Long terme** (>2h) : analyser tendances, planifier maintenance
"""

    def _reponse_defaut(self, question):
        return f"""
### 🤖 Assistant IA Fly-Away

Je n'ai pas bien compris : "{question}"

Essayez : "Latence ?", "Prédictions", "Qualité signal", "Température ?", "Pluie ?", "Recommandations"
"""


# ════════════════════════════════════════════════════════════════════════════
# ÉCRAN AUTHENTIFICATION
# ════════════════════════════════════════════════════════════════════════════

def page_connexion():
    """Page de connexion agent."""
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("## 🛡️ Connexion Agent CRTV")
        st.markdown("---")
        
        matricule = st.text_input("Matricule", placeholder="MAT001")
        password = st.text_input("Mot de passe", type="password", placeholder="••••••")
        lieu = st.selectbox("Lieu Production", LIEUX_DISPONIBLES)
        
        if st.button("🔓 Se Connecter", use_container_width=True, type="primary"):
            if not matricule or not password:
                st.error("Veuillez remplir tous les champs")
                return
            
            succes, message = verifier_authentification(matricule, password)
            
            if succes:
                st.session_state.authenticated = True
                st.session_state.agent_matricule = matricule
                st.session_state.agent_nom = message
                st.session_state.agent_lieu = lieu
                
                agents_db = charger_agents_db()
                if matricule in agents_db:
                    st.session_state.agent_role = agents_db[matricule].get("role", "Opérateur")
                
                st.success(f"✅ Bienvenue {message} !")
                st.rerun()
            else:
                st.error(f"❌ {message}")
        
        st.divider()
        st.caption("**Identifiants de démo** :\n- MAT001 / demo123\n- MAT002 / demo123\n- MAT003 / demo123")


# ════════════════════════════════════════════════════════════════════════════
# APPLICATION PRINCIPALE
# ════════════════════════════════════════════════════════════════════════════

def main():
    # Vérifier authentification
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        page_connexion()
        return
    
    init_global_state()
    
    # Barre latérale — Infos agent + Configuration
    with st.sidebar:
        st.markdown("### 👤 Agent Connecté")
        st.markdown(f"**{st.session_state.agent_nom}**")
        st.caption(f"Matricule : {st.session_state.agent_matricule}")
        st.caption(f"Rôle : {st.session_state.agent_role}")
        st.caption(f"Lieu : {st.session_state.agent_lieu}")
        st.divider()
        
        if st.button("🔓 Déconnexion", use_container_width=True):
            st.session_state.authenticated = False
            st.rerun()
        
        st.divider()
        st.markdown("### ⚙️ Configuration")
        excel_path = st.text_input("Chemin fichier Excel", value=r"C:\Users\HP 1030 G2\OneDrive\Bureau\projetV2\crtv-flyaway-monitor\data\flyaway_log_annuel_2025-1")
        sheet_name = st.text_input("Feuille", value=DEFAULT_SHEET_NAME)
        uploaded_file = st.file_uploader(
            "Ou déposez le fichier Excel ici",
            type=["xlsx", "xls"]
        )
        st.caption(
            "ℹ️ Le tableau se cale sur la dernière donnée du fichier — "
            "pas besoin de changer l'heure système."
        )
    
    # Chargement données
    with st.spinner("⏳ Chargement des données..."):
        df_complet, erreur = charger_donnees(excel_path, sheet_name, uploaded_file)
    
    if erreur:
        st.error(f"❌ {erreur}")
        st.info("Déposez votre fichier Excel dans la barre latérale pour continuer.")
        return
    if df_complet is None or df_complet.empty:
        st.error("❌ Aucune donnée exploitable.")
        return
    
    reference_now = obtenir_reference_temporelle(df_complet)
    df_recent = df_complet[df_complet['timestamp'] >= reference_now - FENETRE_RECENTE]
    if df_recent.empty:
        df_recent = df_complet.head(50)
    
    model_ia = ModeleAIOptimisee()
    model_ia.entrainer_rapide(df_complet)
    
    # En-tête
    col_logo, col_titre = st.columns([1, 5])
    with col_logo:
        if os.path.exists(DEFAULT_LOGO_PATH):
            try:
                st.image(Image.open(DEFAULT_LOGO_PATH), width=80)
            except Exception:
                st.write("🛰️")
        else:
            st.write("🛰️")
    
    with col_titre:
        st.markdown("# 🛰️ CRTV Fly-Away v3.3")
        st.markdown("#### Authentification + Rapports Transmission")
        st.caption(
            f"📍 {st.session_state.agent_lieu} | "
            f"📅 {reference_now.strftime('%Y-%m-%d %H:%M:%S')}"
        )
    
    st.divider()
    
    # Dashboard + Chat
    col_dashboard, col_chat = st.columns([3, 2])
    
    with col_dashboard:
        st.markdown("### 📊 TABLEAU DE BORD")
        
        kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
        with kpi_col1:
            lat = df_recent['latence_totale_ms'].mean()
            st.metric("Latence", f"{lat:.0f}ms", "🔴" if lat > SEUIL_LATENCE_WARN else "🟢")
        with kpi_col2:
            cn0 = df_recent['cn0_dbhz'].mean() if 'cn0_dbhz' in df_recent.columns else np.nan
            st.metric("C/N₀", f"{cn0:.1f}dB", "🟢" if cn0 > 10 else "🔴")
        with kpi_col3:
            qual = df_recent['qualite_signal_pct'].mean() if 'qualite_signal_pct' in df_recent.columns else np.nan
            st.metric("Qualité", f"{qual:.1f}%", "🟢" if qual > 85 else "🔴")
        with kpi_col4:
            temp = df_recent['temperature_hpa_c'].mean() if 'temperature_hpa_c' in df_recent.columns else 0
            st.metric("Temp HPA", f"{temp:.1f}°C", "🟢" if temp < 65 else "🔴")
        
        st.divider()
        
        st.markdown(f"#### Évolution Latence (fenêtre 1h)")
        fig = go.Figure()
        df_sorted = df_recent.sort_values('timestamp')
        fig.add_trace(go.Scatter(
            x=df_sorted['timestamp'], y=df_sorted['latence_totale_ms'],
            mode='lines+markers', name='Latence',
            line=dict(color='#1a472a', width=2), fill='tozeroy'
        ))
        fig.add_hline(y=SEUIL_LATENCE_WARN, line_dash='dash', line_color='orange')
        fig.add_hline(y=SEUIL_LATENCE_CRIT, line_dash='dash', line_color='red')
        fig.update_layout(height=300, hovermode='x unified')
        st.plotly_chart(fig, use_container_width=True)
    
    with col_chat:
        st.markdown("### 🤖 Chat IA")
        
        if "chat_messages" not in st.session_state:
            st.session_state.chat_messages = []
        
        chat_container = st.container(height=400)
        with chat_container:
            for msg in st.session_state.chat_messages:
                with st.chat_message(msg["role"]):
                    st.markdown(msg["content"])
        
        user_input = st.chat_input("Posez une question...")
        if user_input:
            st.session_state.chat_messages.append({"role": "user", "content": user_input})
            chatbot = ChatbotIAPrediction(model_ia, df_recent, df_complet)
            reponse = chatbot.repondre(user_input)
            st.session_state.chat_messages.append({"role": "assistant", "content": reponse})
            st.rerun()
    
    st.divider()
    
    # Onglets
    tab1, tab2, tab3 = st.tabs(["📈 Analyses", "🗂️ Données", "📋 Rapport Transmission"])
    
    with tab1:
        st.markdown("### Analyses Statistiques")
        col_a, col_b = st.columns(2)
        with col_a:
            fig = px.histogram(df_recent, x='latence_totale_ms', nbins=30,
                                color_discrete_sequence=['#1a472a'], title='Distribution Latence')
            st.plotly_chart(fig, use_container_width=True)
        with col_b:
            if 'qualite_signal_pct' in df_recent.columns:
                fig = px.histogram(df_recent, x='qualite_signal_pct', nbins=30,
                                    color_discrete_sequence=['#22c55e'], title='Distribution Qualité')
                st.plotly_chart(fig, use_container_width=True)
    
    with tab2:
        st.markdown("### Données Brutes")
        nb_rows = st.slider("Lignes à afficher", 10, 200, 50)
        st.dataframe(df_recent.head(nb_rows), use_container_width=True, height=300)
        csv = df_recent.to_csv(index=False).encode('utf-8')
        st.download_button(
            "📥 Télécharger CSV", csv,
            f"flyaway_{reference_now.strftime('%Y%m%d_%H%M%S')}.csv"
        )
    
    with tab3:
        st.markdown("### 📄 Rapport de Fin de Transmission")
        st.write(
            "Générez un rapport complet en Excel contenant : résumé, données détaillées, "
            "analyses statistiques et recommandations."
        )
        
        if st.button("✨ Générer Rapport Excel", use_container_width=True, type="primary"):
            with st.spinner("Génération du rapport..."):
                agent_info = {
                    "nom": st.session_state.agent_nom,
                    "matricule": st.session_state.agent_matricule,
                    "lieu": st.session_state.agent_lieu,
                    "role": st.session_state.agent_role
                }
                
                rapport_bytes = generer_rapport_excel(
                    df_recent, df_complet, agent_info, reference_now
                )
                
                st.success("✅ Rapport généré avec succès !")
                st.download_button(
                    "📥 Télécharger Rapport Excel",
                    data=rapport_bytes,
                    file_name=f"Rapport_FlyAway_{reference_now.strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.info("💾 Le rapport contient : résumé exécutif, données détaillées, analyses et recommandations.")


if __name__ == "__main__":
    main()
